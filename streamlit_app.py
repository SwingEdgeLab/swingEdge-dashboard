import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
import requests
import streamlit.components.v1 as components
import warnings
warnings.filterwarnings('ignore')

st.set_page_config(
    page_title="SwingEdge Pro — Member Portal",
    page_icon=None,
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Barlow:ital,wght@0,300;0,400;0,500;0,600;0,700;0,800;0,900;1,300;1,400&family=Barlow+Condensed:wght@300;400;500;600;700;800;900&display=swap');

* { margin: 0; padding: 0; box-sizing: border-box; }

:root {
    --black: #080808;
    --dark: #0e0e0e;
    --card: #141414;
    --card2: #1a1a1a;
    --border: #242424;
    --border2: #2e2e2e;
    --gold: #f5a623;
    --gold2: #ffd066;
    --white: #f0f0f0;
    --muted: #666666;
    --muted2: #999999;
    --green: #22c55e;
    --navy: #1a2035;
}

html, body, [class*="css"] {
    font-family: 'Barlow', sans-serif !important;
    background-color: var(--black) !important;
    color: var(--white) !important;
}
.stApp { background-color: var(--black) !important; }
#MainMenu, footer, header { visibility: hidden; }
section[data-testid="stSidebar"] { display: none !important; }

/* ── NAV ── */
.nav {
    position: fixed; top: 0; left: 0; right: 0;
    background: rgba(8,8,8,0.97);
    border-bottom: 1px solid var(--border);
    padding: 0 2.5rem;
    height: 60px;
    display: flex; align-items: center; justify-content: space-between;
    z-index: 9999;
    backdrop-filter: blur(20px);
}
.nav-brand {
    display: flex; align-items: center; gap: 10px;
}
.nav-logo {
    font-family: 'Barlow', sans-serif;
    font-size: 0.95rem;
    font-weight: 700;
    color: var(--white);
}
.nav-logo span { color: var(--muted2); font-weight: 400; }
.nav-link {
    font-size: 0.82rem; color: var(--muted2); cursor: pointer;
}
.nav-cta {
    background: var(--gold); color: #000;
    font-size: 0.78rem; font-weight: 700;
    padding: 6px 14px; border-radius: 4px; cursor: pointer;
}

/* ── LOGIN ── */
.login-page {
    min-height: 100vh;
    display: flex; align-items: center; justify-content: center;
    background: var(--black);
    padding-top: 60px;
}
.login-card {
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: 4px;
    padding: 3rem 2.5rem;
    width: 100%; max-width: 420px;
    text-align: center;
    position: relative;
    overflow: hidden;
}
.login-card::before {
    content: '';
    position: absolute; top: 0; left: 0; right: 0; height: 2px;
    background: linear-gradient(90deg, transparent, var(--gold), transparent);
}
.login-wolf { font-size: 3rem; margin-bottom: 1rem; }
.login-brand {
    font-family: 'Barlow Condensed', sans-serif;
    font-size: 1.3rem; font-weight: 800;
    letter-spacing: 4px; text-transform: uppercase;
    margin-bottom: 0.25rem;
}
.login-brand span { color: var(--gold); }
.login-tag {
    font-family: 'Barlow Condensed', sans-serif;
    font-size: 0.62rem; letter-spacing: 2px;
    color: var(--muted); text-transform: uppercase;
    margin-bottom: 2.5rem;
}
.login-divider {
    width: 40px; height: 1px;
    background: var(--gold); margin: 0 auto 2rem;
}

/* ── HOME PAGE ── */
.home-hero {
    padding: 6rem 2.5rem 3rem;
    text-align: center;
}
.hero-tag {
    font-family: 'Barlow Condensed', sans-serif;
    font-size: 0.7rem; letter-spacing: 3px;
    color: var(--gold); text-transform: uppercase;
    margin-bottom: 1.5rem;
}
.hero-title {
    font-family: 'Barlow Condensed', sans-serif;
    font-size: 3.5rem; font-weight: 300;
    letter-spacing: 6px; text-transform: uppercase;
    line-height: 1; color: var(--white);
    margin-bottom: 0.5rem;
}
.hero-title strong { font-weight: 800; color: var(--gold); }
.hero-sub {
    font-family: 'Barlow Condensed', sans-serif;
    font-size: 0.75rem; letter-spacing: 4px;
    color: var(--muted); text-transform: uppercase;
    margin-top: 1.5rem;
}

/* ── SCAN GRID ── */
.scans-section {
    padding: 0 2.5rem 4rem;
    max-width: 1400px; margin: 0 auto;
}
.section-label {
    font-family: 'Barlow Condensed', sans-serif;
    font-size: 0.65rem; letter-spacing: 3px;
    color: var(--muted); text-transform: uppercase;
    margin-bottom: 1.5rem;
    padding-bottom: 0.75rem;
    border-bottom: 1px solid var(--border);
}
.home-hero .hero-tag {
    font-family: 'Barlow Condensed', sans-serif;
    font-size: 0.65rem; letter-spacing: 3px;
    color: var(--muted2); text-transform: uppercase;
    margin-bottom: 1.5rem;
}
.home-hero .hero-title {
    font-size: 2.8rem; font-weight: 800;
    line-height: 1.1; color: var(--white);
    margin-bottom: 0.1rem;
}
.home-hero .hero-title em {
    font-style: italic; color: var(--gold);
    font-weight: 800;
}
.home-hero .hero-sub {
    font-size: 0.9rem; color: var(--muted2);
    max-width: 500px; margin: 1.5rem auto 0;
    line-height: 1.6;
}
.features-row {
    display: flex; gap: 1rem;
    padding: 2rem 2.5rem 0;
    max-width: 1400px; margin: 0 auto;
}
.feat-card {
    flex: 1; background: var(--card);
    border: 1px solid var(--border);
    border-radius: 8px; padding: 1.5rem;
}
.feat-icon { font-size: 1.5rem; margin-bottom: 0.75rem; }
.feat-name { font-size: 0.85rem; font-weight: 700; margin-bottom: 0.5rem; }
.feat-desc { font-size: 0.75rem; color: var(--muted2); line-height: 1.5; }
.scan-grid {
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 1px;
    background: var(--border);
    border: 1px solid var(--border);
    border-radius: 4px;
    overflow: hidden;
    margin-bottom: 3rem;
}
.scan-card {
    background: var(--card);
    padding: 1.75rem;
    cursor: pointer;
    transition: background 0.2s;
    position: relative;
}
.scan-card:hover { background: var(--card2); }
.scan-card.active:hover { background: #1a1500; }
.scan-card.active { border-left: 2px solid var(--gold); }
.scan-card.coming { opacity: 0.5; cursor: default; }
.scan-icon { font-size: 1.5rem; margin-bottom: 1rem; }
.scan-name {
    font-family: 'Barlow Condensed', sans-serif;
    font-size: 1rem; font-weight: 700;
    letter-spacing: 1px; text-transform: uppercase;
    color: var(--white); margin-bottom: 0.4rem;
}
.scan-desc {
    font-size: 0.78rem; color: var(--muted2);
    line-height: 1.5; margin-bottom: 1rem;
}
.scan-status {
    font-family: 'Barlow Condensed', sans-serif;
    font-size: 0.6rem; letter-spacing: 2px;
    text-transform: uppercase; font-weight: 700;
}
.scan-status.live { color: var(--gold); }
.scan-status.soon { color: var(--muted); }
.scan-count {
    position: absolute; top: 1.75rem; right: 1.75rem;
    font-family: 'Barlow Condensed', sans-serif;
    font-size: 1.5rem; font-weight: 800;
    color: var(--gold); opacity: 0.3;
}
.scan-arrow {
    position: absolute; bottom: 1.75rem; right: 1.75rem;
    color: var(--gold); font-size: 1rem; opacity: 0;
    transition: opacity 0.2s;
}
.scan-card.active:hover .scan-arrow { opacity: 1; }

/* ── DASHBOARD ── */
.dash-header {
    padding: 5rem 2.5rem 1.5rem;
}
.dash-back {
    font-family: 'Barlow Condensed', sans-serif;
    font-size: 0.7rem; letter-spacing: 2px;
    color: var(--muted); text-transform: uppercase;
    cursor: pointer; margin-bottom: 1.5rem;
    display: inline-block;
}
.dash-title {
    font-family: 'Barlow Condensed', sans-serif;
    font-size: 2rem; font-weight: 800;
    letter-spacing: 2px; text-transform: uppercase;
    color: var(--white); margin-bottom: 0.25rem;
}
.dash-title span { color: var(--gold); }
.dash-meta {
    font-size: 0.78rem; color: var(--muted2);
    font-family: 'Barlow Condensed', sans-serif;
    letter-spacing: 1px;
}

.metrics-strip {
    display: grid;
    grid-template-columns: repeat(5, 1fr);
    gap: 1px; background: var(--border);
    margin: 0 2.5rem 1px;
    border: 1px solid var(--border);
    border-radius: 4px 4px 0 0; overflow: hidden;
}
.metric {
    background: var(--card);
    padding: 1.25rem 1rem;
    text-align: center;
}
.metric.highlight { background: #110f00; }
.metric-n {
    font-family: 'Barlow Condensed', sans-serif;
    font-size: 2.5rem; font-weight: 900;
    color: var(--gold); line-height: 1;
}
.metric-n.gold2 { color: var(--gold2); }
.metric-l {
    font-family: 'Barlow Condensed', sans-serif;
    font-size: 0.6rem; letter-spacing: 2px;
    color: var(--muted); text-transform: uppercase;
    margin-top: 4px; font-weight: 600;
}

/* ── TABS ── */
.stTabs [data-baseweb="tab-list"] {
    background: var(--card) !important;
    border-bottom: 1px solid var(--gold) !important;
    gap: 0 !important; padding: 0 !important;
    border-radius: 0 !important;
}
.stTabs [data-baseweb="tab"] {
    background: transparent !important;
    color: var(--muted2) !important;
    font-family: 'Barlow Condensed', sans-serif !important;
    font-weight: 700 !important;
    font-size: 0.78rem !important;
    letter-spacing: 1.5px !important;
    text-transform: uppercase !important;
    padding: 0.75rem 1.5rem !important;
    border-radius: 0 !important;
}
.stTabs [aria-selected="true"] {
    background: var(--gold) !important;
    color: #000 !important;
}
.stTabs [data-baseweb="tab-panel"] {
    background: var(--card) !important;
    border: 1px solid var(--border) !important;
    border-top: none !important;
    padding: 1.25rem !important;
    border-radius: 0 0 4px 4px !important;
}

/* ── INPUTS / BUTTONS ── */
.stTextInput input {
    background: var(--dark) !important;
    border: 1px solid var(--border2) !important;
    color: var(--white) !important;
    border-radius: 3px !important;
    font-family: 'Barlow', sans-serif !important;
    font-size: 0.85rem !important;
}
.stTextInput input:focus { border-color: var(--gold) !important; }
.stButton button {
    background: var(--gold) !important;
    color: #000 !important;
    font-family: 'Barlow Condensed', sans-serif !important;
    font-weight: 800 !important;
    letter-spacing: 2px !important;
    text-transform: uppercase !important;
    border: none !important;
    border-radius: 3px !important;
}
.stDownloadButton button {
    background: transparent !important;
    color: var(--gold) !important;
    border: 1px solid var(--gold) !important;
    font-family: 'Barlow Condensed', sans-serif !important;
    font-weight: 700 !important;
    letter-spacing: 1px !important;
    border-radius: 3px !important;
}

/* Fullscreen fix */
[data-testid="stFullScreenFrame"] {
    background: #080808 !important;
}
.fullscreen-wrapper {
    background: #080808 !important;
}
iframe[title="st.dataframe"] {
    background: #080808 !important;
}
.footer {
    text-align: center;
    font-family: 'Barlow Condensed', sans-serif;
    font-size: 0.6rem; letter-spacing: 2px;
    color: var(--border2); text-transform: uppercase;
    padding: 2rem 0 3rem;
}
</style>
""", unsafe_allow_html=True)

# ── PASSWORD ──────────────────────────────────────────────────
PASSWORD = "SwingEdge@2026"

if "auth" not in st.session_state:
    st.session_state.auth = False
if "page" not in st.session_state:
    st.session_state.page = "home"

# ── NAV ──────────────────────────────────────────────────────
st.markdown("""
<div class="nav">
  <div class="nav-brand">
    <span style="color:#f5a623; font-size:0.6rem;">●</span>
    <span class="nav-logo">SwingEdge<span>Pro</span></span>
  </div>
  <div style="display:flex; gap:2rem; align-items:center;">
    <span class="nav-link">Indicator</span>
    <span class="nav-link">Why It Works</span>
    <span class="nav-link">The Proof</span>
  </div>
  <div style="display:flex; gap:1rem; align-items:center;">
    <span class="nav-link">Log in</span>
    <span class="nav-cta">Member Portal</span>
  </div>
</div>
""", unsafe_allow_html=True)

# ── LOGIN ─────────────────────────────────────────────────────
if not st.session_state.auth:
    st.markdown('<div style="height:12vh"></div>', unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 1.1, 1])
    with col2:
        st.markdown("""
        <div class="login-card">
          <div style="font-size:0.6rem; color:#f5a623; letter-spacing:2px; margin-bottom:0.5rem;">● SWINGEDGEPRO.IN</div>
          <div class="login-brand">SwingEdge<span>Pro</span></div>
          <div class="login-tag">Member Access · Daily Scans</div>
          <div class="login-divider"></div>
        </div>
        """, unsafe_allow_html=True)
        pwd = st.text_input("", type="password", placeholder="Enter member password...", label_visibility="collapsed")
        if st.button("ACCESS PORTAL →", use_container_width=True):
            if pwd == PASSWORD:
                st.session_state.auth = True
                st.rerun()
            else:
                st.error("Incorrect password.")
        st.markdown("""
        <div style="text-align:center; margin-top:1.5rem;">
          <div style="font-family:'Barlow Condensed',sans-serif; font-size:0.6rem; letter-spacing:2px; color:#444; text-transform:uppercase;">
            ALIGN WITH MOMENTUM · EXECUTE WITH AN EDGE
          </div>
        </div>
        """, unsafe_allow_html=True)
    st.stop()

# ── SCAN CONFIG ───────────────────────────────────────────────
SCANS = [
    {
        "key": "indicator",
        "icon": "⚡",
        "name": "SwingEdge Pro Ultimate",
        "desc": "6-in-1 TradingView indicator. One glance to know if a stock deserves your watchlist.",
        "status": "live",
    },
    {
        "key": "bottom_bounce",
        "icon": "📊",
        "name": "Bottom Bounce",
        "desc": "EMA crossover recovery — stocks emerging from stage 1 base with volume confirmation",
        "status": "live",
        "sheets": {
            'T0': '👀 T0 — Pre-Recovery',
            'T1': '🌱 T1 — 10W Cross 20W',
            'T2': '🔥 T2 — 10W Cross 40W',
            'T3': '🚀 T3 — 20W Cross 40W',
            'COMBO': '⚡ COMBO T2+T3',
        },
        "tab_labels": lambda c: [f"T0 ({c['T0']})", f"T1 ({c['T1']})", f"T2 ({c['T2']})", f"T3 ({c['T3']})", f"COMBO ({c['COMBO']})"],
        "desc_map": {
            'T0': 'Pre-Recovery — Gap narrowing, EMAs still inverted',
            'T1': 'Early Cross — 10W crossed above 20W',
            'T2': 'Momentum — 10W crossed above 40W',
            'T3': 'Trend Restored — 20W crossed above 40W',
            'COMBO': 'High Conviction — T2 + T3 both fired this week',
        }
    },
    {"key": "rs_explosion", "icon": "🚀", "name": "RS Explosion", "desc": "Relative strength breakout — stocks showing institutional accumulation with RS surge above 70", "status": "live"},
    {"key": "l1_structural", "icon": "🏆", "name": "L1 Structural Quality", "desc": "ELITE / PRIME / STRONG / WATCH — 9-condition Minervini SEPA scoring across the full universe", "status": "live"},
    {"key": "l2_institutional", "icon": "🏛️", "name": "L2 Institutional Volume", "desc": "CLIMAX / TIER 1 / TIER 2 / WATCH — institutional volume signature detection", "status": "live"},
    {"key": "rs_screener", "icon": "📡", "name": "RS Screener", "desc": "Relative strength ranked universe — RS90+ elite, structure hunt 65-90, all EMAs aligned", "status": "live"},
    {"key": "ep_scanner", "icon": "⚡", "name": "EP Scanner", "desc": "Episodic Pivot — EXPLOSIVE / STRONG / VALID — news-driven volume surge setups", "status": "live"},
    {"key": "htf_scanner", "icon": "🎯", "name": "HTF Scanner", "desc": "High Tight Flag patterns — parabolic base with tight consolidation before next leg", "status": "live"},
    {"key": "nr_3wtc", "icon": "📐", "name": "3WTC / NR Scanner", "desc": "3-Week Tight Close and Narrow Range patterns — coiling before breakout", "status": "live"},
    {"key": "weekly_scanner", "icon": "📅", "name": "Weekly Scanner", "desc": "Power Breakouts / Valid Breakouts / 3WTC / Inside Week / Mini Coil / NR7 — weekly setups", "status": "live"},
    {"key": "ultrapro_weekly", "icon": "💎", "name": "UltraPro Weekly", "desc": "ULTRA PRIME / STRONG / WATCH — highest conviction weekly momentum setups", "status": "live"},
    {"key": "vdu_stage2", "icon": "📶", "name": "VDU Stage 2", "desc": "Volume-Demand-Urgency Stage 2 — Elite / 3WTC / Established / Transition setups", "status": "live"},
    {"key": "resilient_stack", "icon": "🛡️", "name": "Resilient Stack Reversal", "desc": "Stocks holding EMA stack during market weakness — high relative strength in downturns", "status": "live"},
    {"key": "market_health", "icon": "🩺", "name": "Market Health Dashboard", "desc": "9-9 Watchlist / 8-9 Monitor / RS Leaders / Rally Quality / N500 Surge — daily breadth", "status": "live"},
    {"key": "rs_outperformer", "icon": "🥇", "name": "RS Outperformer", "desc": "Stocks outperforming Nifty since a chosen date — Elite / Strong / Mild RS grades", "status": "live"},
    {"key": "52w_high", "icon": "🏔️", "name": "52W High Breakout", "desc": "Stocks breaking out of multi-week bases near 52-week highs with volume expansion", "status": "live"},
    {"key": "accumulation", "icon": "🏦", "name": "Accumulation", "desc": "Silent institutional buying — 3+ up days vs down days with 1.5x volume ratio", "status": "live"},
    {"key": "ath_scanner", "icon": "⚡", "name": "ATH Scanner", "desc": "Stocks within 5% of all-time highs with Stage 2 structure and RS ≥ 70", "status": "live"},
    {"key": "ema200_cross", "icon": "📈", "name": "200 EMA Cross", "desc": "Fresh crossovers above the 200 EMA with RS confirmation and volume surge", "status": "live"},
    {"key": "deep_base", "icon": "🔭", "name": "Deep Base Recovery", "desc": "Long consolidation breakout — stocks recovering from extended Stage 1 bases", "status": "live"},
    {"key": "50ema_shakeout", "icon": "💥", "name": "50 EMA Shakeout", "desc": "Stage 2 leaders retesting 50 EMA — high-conviction pullback entry setups", "status": "live"},
    {"key": "market_verdict", "icon": "📡", "name": "Market Verdict", "desc": "Enter today's breadth readings — get your market health score and position sizing guidance", "status": "live"},
    {"key": "score_guide", "icon": "📋", "name": "Score Guide", "desc": "Reference guide for SwingEdge Pro composite scores and what each level means", "status": "live"},
]

# ── GENERIC SCANNER DASHBOARD HELPER ─────────────────────────
def render_generic_scanner(page_key, title, icon, accent_color, sheet_config, file_prefix, desc_map=None):
    """
    Generic dashboard renderer for scanners that push Excel files to GitHub.
    - Downloads the raw .xlsx bytes once from GitHub
    - Shows a preview dataframe per tab
    - Download button serves the ORIGINAL styled Excel file as-is
    """
    if st.button("← Back to Portal", key=f"back_{page_key}"):
        st.session_state.page = "home"
        st.rerun()

    @st.cache_data(ttl=300)
    def load_scanner(prefix):
        try:
            api = requests.get(
                "https://api.github.com/repos/SwingEdgeLab/swingEdge-dashboard/contents/"
            ).json()
            files = sorted(
                [f['name'] for f in api
                 if isinstance(f, dict) and f['name'].startswith(prefix) and f['name'].endswith('.xlsx')],
                reverse=True
            )
            if not files:
                return None, None, "No file found", "—", "—"
            fname = files[0]
            url   = f"https://github.com/SwingEdgeLab/swingEdge-dashboard/raw/main/{fname}"
            r     = requests.get(url, headers={"Accept": "application/octet-stream"})
            raw_bytes = r.content                          # ← keep original bytes
            wb        = openpyxl.load_workbook(BytesIO(raw_bytes), data_only=True)

            # ── Read Scan Info sheet ──
            scan_date, run_time = "—", "—"
            for info_sn in ("📅 Scan Info", "Scan Info"):
                if info_sn in wb.sheetnames:
                    ws_info = wb[info_sn]
                    for row in ws_info.iter_rows(min_row=2, values_only=True):
                        if not row or len(row) < 2 or not row[0]:
                            continue
                        k = str(row[0]).strip()
                        v = str(row[1]).strip() if row[1] else ""
                        if k == "Scan Date":
                            scan_date = v
                        elif k == "Run Time":
                            run_time = v
                        elif k in ("Scan Run", "Last Run", "Run Date & Time"):
                            parts = v.split("  ")
                            scan_date = parts[0].strip() if parts else v
                            run_time  = parts[1].strip() if len(parts) > 1 else "—"
                    break

            # ── Load preview dataframes per tab ──
            data = {}
            for sheet_prefix, _ in sheet_config:
                matched = next((sn for sn in wb.sheetnames if sn.startswith(sheet_prefix)), None)
                if matched:
                    ws   = wb[matched]
                    rows = list(ws.iter_rows(values_only=True))
                    if len(rows) < 2:
                        data[sheet_prefix] = pd.DataFrame()
                        continue
                    hdrs = [str(h) if h is not None else f"_c{i}" for i, h in enumerate(rows[0])]
                    df   = pd.DataFrame(rows[1:], columns=hdrs)
                    if hdrs[0].startswith("_c") or hdrs[0] == "Unnamed: 0":
                        df = df.iloc[:, 1:] if len(df.columns) > 1 else df
                    if "Symbol" in df.columns:
                        df = df[df["Symbol"].notna() & (df["Symbol"].astype(str).str.strip() != "")]
                    data[sheet_prefix] = df
                else:
                    data[sheet_prefix] = pd.DataFrame()

            return data, raw_bytes, fname, scan_date, run_time
        except Exception as e:
            return None, None, "Error", str(e), "—"

    with st.spinner("Loading scan data..."):
        result = load_scanner(file_prefix)

    data, raw_bytes, fname, scan_date, run_time = result

    if data is None:
        st.error(f"Could not load scan data ({fname}). Ensure the file is pushed to GitHub.")
        st.stop()

    counts = {sp: len(data.get(sp, pd.DataFrame())) for sp, _ in sheet_config}
    total  = sum(counts.values())

    pills_html = "".join(
        f'<span style="color:#888;">{lbl.split("—")[0].strip()} '
        f'<b style="color:{accent_color};">{counts[sp]}</b></span>'
        for sp, lbl in sheet_config
    )

    # ── Header row with title + Download Original button ──
    h1, h2 = st.columns([3, 1])
    with h1:
        st.markdown(f"""
        <div style="padding:4.5rem 0 0.5rem; display:flex; align-items:center; flex-wrap:wrap; gap:0.5rem;">
          <div>
            <span style="font-family:'Barlow Condensed',sans-serif; font-size:1.1rem;
                         font-weight:800; letter-spacing:2px; text-transform:uppercase;
                         color:#f0f0f0;">{icon} <span style="color:{accent_color};">{title}</span></span>
            <span style="font-size:0.72rem; color:#666; margin-left:1rem;
                         font-family:'Barlow Condensed',sans-serif; letter-spacing:1px;">
              {scan_date} · {run_time} · 1304 stocks
            </span>
          </div>
          <div style="display:flex; gap:1.5rem; font-family:'Barlow Condensed',sans-serif;
                      font-size:0.72rem; letter-spacing:1px; flex-wrap:wrap; margin-top:0.4rem;">
            {pills_html}
            <span style="color:#888;">TOTAL <b style="color:{accent_color};">{total}</b></span>
          </div>
        </div>
        """, unsafe_allow_html=True)
    with h2:
        st.markdown('<div style="padding-top:4.5rem;"></div>', unsafe_allow_html=True)
        st.download_button(
            label="⬇ Download Original Excel",
            data=raw_bytes,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"dl_orig_{page_key}",
            use_container_width=True,
        )

    # ── Tabs ──
    tab_labels = [f"{lbl.split('—')[0].strip()}  ({counts[sp]})" for sp, lbl in sheet_config]
    tabs = st.tabs(tab_labels)

    for tab_idx, (tab, (sheet_prefix, label)) in enumerate(zip(tabs, sheet_config)):
        with tab:
            df = data.get(sheet_prefix, pd.DataFrame())
            if df.empty:
                st.info("No signals for this category today.")
            else:
                if desc_map and sheet_prefix in desc_map:
                    st.caption(f"_{desc_map[sheet_prefix]}_")
                search = st.text_input(
                    "", key=f"s_{page_key}_{tab_idx}", placeholder="Search symbol..."
                )
                if search and "Symbol" in df.columns:
                    df = df[df["Symbol"].astype(str).str.contains(search.upper(), na=False)]
                st.dataframe(df, height=min(900, max(400, len(df) * 38)),
                             hide_index=True, use_container_width=True)

    st.markdown(
        '<div class="footer">SwingEdgePro.in · Confidential · Member Use Only · © 2026</div>',
        unsafe_allow_html=True
    )


# ── HOME PAGE ─────────────────────────────────────────────────
if st.session_state.page == "home":
    st.markdown("""
    <div class="home-hero">
      <div class="hero-tag">RECLAIM YOUR EDGE</div>
      <div class="hero-title">Designed for the <em>serious</em></div>
      <div class="hero-title">momentum trader.</div>
      <div class="hero-sub">We do the heavy lifting so you can focus on what matters — finding the right stocks at the right time.</div>
    </div>

    <div class="features-row">
      <div class="feat-card">
        <div class="feat-icon">🚫</div>
        <div class="feat-name">Eliminate Dead Stocks</div>
        <div class="feat-desc">Only stocks showing real strength, momentum and volume qualify. No more buying into weakness.</div>
      </div>
      <div class="feat-card">
        <div class="feat-icon">🔗</div>
        <div class="feat-name">Connect the Signals</div>
        <div class="feat-desc">6 powerful metrics combined into one indicator. EMA, RS, Volume, ADR% — all in one glance.</div>
      </div>
      <div class="feat-card">
        <div class="feat-icon">💡</div>
        <div class="feat-name">Instant Clarity</div>
        <div class="feat-desc">Know immediately if a stock deserves your attention. No more second-guessing or manual checks.</div>
      </div>
      <div class="feat-card">
        <div class="feat-icon">🎯</div>
        <div class="feat-name">Impulse + Tightness</div>
        <div class="feat-desc">Built around the exact setup that produces explosive moves — coiling before the breakout.</div>
      </div>
      <div class="feat-card">
        <div class="feat-icon">⏱️</div>
        <div class="feat-name">Save Hours Daily</div>
        <div class="feat-desc">What used to take hours of manual scanning now takes seconds. Your edge, automated.</div>
      </div>
    </div>

    <div style="height:3rem;"></div>
    """, unsafe_allow_html=True)

    # ── MEMBER LOGIN SECTION ──
    st.markdown('<div id="member-portal" style="padding: 3rem 2.5rem; text-align:center; border-top: 1px solid #242424;">', unsafe_allow_html=True)
    st.markdown("""
    <div style="font-family:'Barlow Condensed',sans-serif; font-size:0.65rem; letter-spacing:3px; color:#888; text-transform:uppercase; margin-bottom:1rem;">— Member Portal</div>
    <div style="font-family:'Barlow',sans-serif; font-size:1.5rem; font-weight:800; color:#f0f0f0; margin-bottom:0.5rem;">Access Your Daily Scans</div>
    <div style="font-size:0.82rem; color:#888; margin-bottom:2rem;">Enter your member password to view today's scan results</div>
    """, unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 1, 1])
    with col2:
        pwd = st.text_input("", type="password", placeholder="Enter member password...", label_visibility="collapsed", key="home_pwd")
        if st.button("ACCESS PORTAL →", use_container_width=True, key="home_login"):
            if pwd == PASSWORD:
                st.session_state.auth = True
                st.rerun()
            else:
                st.error("Incorrect password.")
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div style="padding: 0 1rem;">', unsafe_allow_html=True)
    st.markdown('<div class="section-label">— Scan Intelligence Suite</div>', unsafe_allow_html=True)

    # Render scan cards in 3-column grid using st.columns
    rows = [SCANS[i:i+3] for i in range(0, len(SCANS), 3)]
    for row in rows:
        cols = st.columns(3)
        for col, scan in zip(cols, row):
            with col:
                is_live = scan['status'] == 'live'
                border = "border-left: 3px solid #f5a623;" if is_live else "opacity: 0.5;"
                status = "● LIVE" if is_live else "○ COMING SOON"
                status_color = "#f5a623" if is_live else "#666"
                st.markdown(f"""
                <div style="background:#141414; {border} padding:1.5rem; border-radius:4px; margin-bottom:1px; min-height:180px;">
                  <div style="font-size:1.5rem; margin-bottom:0.75rem;">{scan['icon']}</div>
                  <div style="font-family:'Barlow Condensed',sans-serif; font-size:1rem; font-weight:800; letter-spacing:2px; text-transform:uppercase; color:#f0f0f0; margin-bottom:0.4rem;">{scan['name']}</div>
                  <div style="font-size:0.75rem; color:#888; line-height:1.5; margin-bottom:1rem;">{scan['desc']}</div>
                  <div style="font-family:'Barlow Condensed',sans-serif; font-size:0.6rem; letter-spacing:2px; font-weight:700; color:{status_color};">{status}</div>
                </div>
                """, unsafe_allow_html=True)
                if is_live:
                    if st.button(f"Open {scan['name']} →", key=f"open_{scan['key']}", use_container_width=True):
                        st.session_state.page = scan['key']
                        st.rerun()

    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown('<div class="footer">SwingEdgePro.in · Confidential · Member Use Only · © 2026</div>', unsafe_allow_html=True)

# ── INDICATOR PAGE ───────────────────────────────────────────
elif st.session_state.page == "indicator":

    if st.button("← Back to Portal", key="back_ind"):
        st.session_state.page = "home"
        st.rerun()

    st.markdown("""
    <div class="dash-header">
      <div style="font-family:'Barlow Condensed',sans-serif; font-size:0.65rem; letter-spacing:3px; color:#f5a623; text-transform:uppercase; margin-bottom:0.75rem;">+ SwingEdge Pro Indicator</div>
      <div style="font-size:2.5rem; font-weight:800; color:#f0f0f0; line-height:1.1; margin-bottom:0.5rem;">One indicator.<br><em style="color:#f5a623; font-style:italic;">Complete stock</em><br>intelligence.</div>
      <div style="font-size:0.85rem; color:#888; max-width:500px; margin-bottom:1.5rem; line-height:1.6;">Built on TradingView. Source protected. Every metric you need to qualify a stock for your watchlist — merged into one powerful, private indicator.</div>
      <div style="background:#1a1a1a; border-left:3px solid #f5a623; padding:1rem 1.25rem; border-radius:4px; max-width:500px; margin-bottom:2rem;">
        <em style="font-size:0.85rem; color:#ccc;">"The only indicator that tells you in one glance whether a stock deserves to be on your watchlist or not."</em>
      </div>
    </div>
    """, unsafe_allow_html=True)

    col1, col2 = st.columns([1, 1])
    with col2:
        st.markdown("""
        <div style="background:#141414; border:1px solid #2e2e2e; border-radius:12px; padding:2rem; margin-top:2rem;">
          <div style="display:flex; align-items:center; gap:0.75rem; margin-bottom:0.25rem;">
            <div style="width:4px; height:1.2rem; background:#f5a623; border-radius:2px;"></div>
            <div style="font-weight:800; font-size:1rem; color:#f0f0f0;">SwingEdge Pro Ultimate</div>
          </div>
          <div style="font-family:'Barlow Condensed',sans-serif; font-size:0.6rem; letter-spacing:2px; color:#f5a623; text-transform:uppercase; margin-bottom:1.5rem; padding-left:1rem;">+ 6-IN-1 INDICATOR</div>
          <div style="display:flex; flex-direction:column; gap:0;">
            %s
          </div>
          <div style="margin-top:1.5rem; padding-top:1rem; border-top:1px solid #242424; display:flex; justify-content:space-between; align-items:center;">
            <div style="font-size:0.72rem; color:#666;">🔒 Source Protected · Not for resale</div>
            <div style="font-size:0.78rem; font-weight:700; color:#f5a623;">SwingEdge Pro</div>
          </div>
        </div>
        """ % "".join([
            f"""<div style="display:flex; justify-content:space-between; align-items:center; padding:0.85rem 0; border-bottom:1px solid #1e1e1e;">
              <div style="display:flex; gap:0.75rem; align-items:center;">
                <span style="font-family:'Barlow Condensed',sans-serif; font-size:0.7rem; color:#555; font-weight:600;">0{i+1}</span>
                <span style="font-size:0.88rem; font-weight:600; color:#f0f0f0;">{name}</span>
              </div>
              <span style="color:#22c55e; font-size:0.8rem;">✓</span>
            </div>"""
            for i, name in enumerate(["Trend Levels", "Pace Quality", "Recovery %", "Relative Volume", "Money Flow Quality", "Strength Score"])
        ]), unsafe_allow_html=True)

    st.markdown('<div class="footer">SwingEdgePro.in · Confidential · Member Use Only · © 2026</div>', unsafe_allow_html=True)

# ── MARKET VERDICT ────────────────────────────────────────────
elif st.session_state.page == "market_verdict":
    if st.button("← Back to Portal", key="back_mv"):
        st.session_state.page = "home"
        st.rerun()
    st.components.v1.html("""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>SwingEdge Pro - Market Verdict</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body {
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
    background: #0b1220;
    color: #e2e8f0;
    min-height: 100vh;
    padding: 2rem 1rem;
  }
  .container { max-width: 720px; margin: 0 auto; }
  .header { margin-bottom: 2rem; }
  .header h1 { font-size: 22px; font-weight: 500; color: #f59e0b; margin-bottom: 4px; }
  .header p { font-size: 13px; color: #64748b; }
  .inputs { display: grid; grid-template-columns: repeat(auto-fit, minmax(130px, 1fr)); gap: 12px; margin-bottom: 1.25rem; }
  .input-card {
    background: #111827;
    border: 0.5px solid #1e2d45;
    border-radius: 10px;
    padding: 1rem;
  }
  .input-card label {
    font-size: 11px;
    color: #64748b;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    display: block;
    margin-bottom: 8px;
  }
  .input-card input {
    width: 100%;
    font-size: 22px;
    font-weight: 500;
    background: transparent;
    border: none;
    border-bottom: 1.5px solid #1e2d45;
    padding: 4px 0;
    color: #e2e8f0;
    outline: none;
  }
  .input-card input:focus { border-bottom-color: #f59e0b; }
  .btn {
    width: 100%;
    padding: 14px;
    font-size: 15px;
    font-weight: 500;
    background: #1e2d45;
    color: #e2e8f0;
    border: 0.5px solid #334155;
    border-radius: 10px;
    cursor: pointer;
    margin-bottom: 1.5rem;
    transition: background 0.2s;
  }
  .btn:hover { background: #253548; }
  .verdict-card {
    border-radius: 12px;
    padding: 1.5rem;
    margin-bottom: 1rem;
    border: 0.5px solid #1e2d45;
  }
  .verdict-top { display: flex; justify-content: space-between; align-items: flex-start; flex-wrap: wrap; gap: 12px; margin-bottom: 1rem; }
  .verdict-label { font-size: 11px; color: #64748b; text-transform: uppercase; letter-spacing: 0.06em; margin-bottom: 4px; }
  .verdict-name { font-size: 28px; font-weight: 500; }
  .verdict-sub { font-size: 13px; color: #94a3b8; margin-top: 4px; }
  .score-num { font-size: 38px; font-weight: 500; }
  .score-denom { font-size: 12px; color: #64748b; }
  .progress-track { height: 8px; background: #1e2d45; border-radius: 4px; overflow: hidden; }
  .progress-fill { height: 100%; border-radius: 4px; transition: width 0.5s; }
  .section {
    background: #111827;
    border: 0.5px solid #1e2d45;
    border-radius: 12px;
    padding: 1.25rem;
    margin-bottom: 1rem;
  }
  .section-title { font-size: 13px; font-weight: 500; margin-bottom: 12px; color: #e2e8f0; }
  .sizing-text { font-size: 15px; line-height: 1.6; color: #e2e8f0; }
  .signal-row { display: flex; align-items: center; justify-content: space-between; gap: 8px; padding: 6px 0; border-bottom: 0.5px solid #1e2d45; }
  .signal-row:last-child { border-bottom: none; }
  .signal-left { display: flex; align-items: center; gap: 8px; }
  .dot { width: 8px; height: 8px; border-radius: 50%; flex-shrink: 0; }
  .signal-name { font-size: 13px; color: #e2e8f0; }
  .signal-right { text-align: right; }
  .signal-val { font-size: 13px; font-weight: 500; }
  .signal-note { font-size: 12px; color: #64748b; margin-left: 8px; }
  .alert-text { font-size: 14px; line-height: 1.7; color: #e2e8f0; }
  .footer { margin-top: 2rem; font-size: 11px; color: #334155; text-align: center; }
  #output { display: none; }
</style>
</head>
<body>
<div class="container">

  <div class="header">
    <h1>SwingEdge Pro — Market Verdict</h1>
    <p>Enter today's breadth readings to get your market health verdict and position sizing guidance.</p>
  </div>

  <div class="inputs">
    <div class="input-card">
      <label>9/9 Minervini count</label>
      <input type="number" id="m9" min="0" max="500" value="256">
    </div>
    <div class="input-card">
      <label>% above 200 EMA</label>
      <input type="number" id="e200" min="0" max="100" value="46">
    </div>
    <div class="input-card">
      <label>% above 50 EMA</label>
      <input type="number" id="e50" min="0" max="100" value="73">
    </div>
    <div class="input-card">
      <label>% above 20 EMA</label>
      <input type="number" id="e20" min="0" max="100" value="69">
    </div>
    <div class="input-card">
      <label>% above 10 EMA</label>
      <input type="number" id="e10" min="0" max="100" value="53">
    </div>
  </div>

  <button class="btn" onclick="calculate()">Calculate market verdict</button>

  <div id="output">

    <div class="verdict-card" id="verdict-card">
      <div class="verdict-top">
        <div>
          <div class="verdict-label">Market verdict</div>
          <div class="verdict-name" id="verdict-name"></div>
          <div class="verdict-sub" id="verdict-sub"></div>
        </div>
        <div style="text-align:right">
          <div class="verdict-label">Composite score</div>
          <div class="score-num" id="score-num"></div>
          <div class="score-denom">out of 100</div>
        </div>
      </div>
      <div class="progress-track">
        <div class="progress-fill" id="progress-fill"></div>
      </div>
    </div>

    <div class="section">
      <div class="section-title">Position sizing rule</div>
      <div class="sizing-text" id="sizing-text"></div>
    </div>

    <div class="section">
      <div class="section-title">Signal breakdown</div>
      <div id="signals"></div>
    </div>

    <div class="section">
      <div class="section-title">Key alert</div>
      <div class="alert-text" id="alert-text"></div>
    </div>

  </div>

  <div class="footer">SwingEdge Pro &mdash; Your Edge. Every Week.</div>

</div>

<script>
function calculate() {
  var m9   = parseFloat(document.getElementById('m9').value)   || 0;
  var e200 = parseFloat(document.getElementById('e200').value) || 0;
  var e50  = parseFloat(document.getElementById('e50').value)  || 0;
  var e20  = parseFloat(document.getElementById('e20').value)  || 0;
  var e10  = parseFloat(document.getElementById('e10').value)  || 0;

  var score = Math.round(
    Math.min(m9 / 300, 1) * 35 +
    (e200 / 100) * 30 +
    (e50  / 100) * 20 +
    (e20  / 100) * 10 +
    (e10  / 100) * 5
  );

  var bullConfirmed = e200 >= 50;
  var stLeader      = m9 >= 150;
  var stWeak        = e10 < 40;
  var divergence    = (e50 >= 70 || e20 >= 70) && e200 < 50;

  var verdict, sub, color, bg, sizing;

  if (score >= 75 && bullConfirmed) {
    verdict = "Full bull";             color = "#22c55e"; bg = "rgba(34,197,94,0.08)";
    sub     = "Broad participation confirmed. Leaders strong.";
    sizing  = "Full position sizing. All setups eligible — ELITE, PRIME, STRONG.";
  } else if (score >= 60 && bullConfirmed) {
    verdict = "Healthy bull";          color = "#4ade80"; bg = "rgba(74,222,128,0.08)";
    sub     = "Bull confirmed. Some short-term softness.";
    sizing  = "Normal sizing — 75 to 100% of planned position. ELITE and PRIME setups.";
  } else if (score >= 50 && !bullConfirmed && stLeader) {
    verdict = "Selective bull";        color = "#f59e0b"; bg = "rgba(245,158,11,0.08)";
    sub     = "Leaders strong. Broad market not yet confirmed.";
    sizing  = "Half size — ELITE setups only. Wait for % above 200 EMA to cross 50%.";
  } else if (score >= 40) {
    verdict = "Choppy / recovering";   color = "#fb923c"; bg = "rgba(251,146,60,0.08)";
    sub     = "Mixed signals. Be very selective.";
    sizing  = "25 to 50% size — ELITE only. Tight stops. Reduce on any failed breakout.";
  } else if (score >= 25) {
    verdict = "Weak / correction";     color = "#ef4444"; bg = "rgba(239,68,68,0.08)";
    sub     = "Avoid new longs. Protect capital.";
    sizing  = "No new trades. Exit weak positions. Raise cash.";
  } else {
    verdict = "Bear / avoid";          color = "#dc2626"; bg = "rgba(220,38,38,0.10)";
    sub     = "Cash is a position. Stay out.";
    sizing  = "100% cash. No longs under any circumstance.";
  }

  var vc = document.getElementById('verdict-card');
  vc.style.background = bg;
  vc.style.borderColor = color + '33';

  document.getElementById('verdict-name').textContent = verdict;
  document.getElementById('verdict-name').style.color = color;
  document.getElementById('verdict-sub').textContent  = sub;
  document.getElementById('score-num').textContent    = score;
  document.getElementById('score-num').style.color    = color;
  document.getElementById('progress-fill').style.width      = score + '%';
  document.getElementById('progress-fill').style.background = color;
  document.getElementById('sizing-text').textContent  = sizing;

  var signals = [
    { label: '9/9 Minervini count', val: m9.toFixed(0),        ok: m9 >= 150,   warn: m9 >= 80,   note: m9 >= 200 ? 'Super bull' : m9 >= 150 ? 'Strong leaders' : m9 >= 80 ? 'Moderate' : 'Weak leadership' },
    { label: '% above 200 EMA',     val: e200.toFixed(1) + '%', ok: e200 >= 50,  warn: e200 >= 35, note: e200 >= 50 ? 'Bull confirmed' : e200 >= 35 ? 'Near threshold' : 'Not confirmed' },
    { label: '% above 50 EMA',      val: e50.toFixed(1) + '%',  ok: e50 >= 60,   warn: e50 >= 40,  note: e50 >= 80 ? 'Overbought' : e50 >= 60 ? 'Healthy' : 'Weak' },
    { label: '% above 20 EMA',      val: e20.toFixed(1) + '%',  ok: e20 >= 55,   warn: e20 >= 35,  note: e20 >= 80 ? 'Overbought ST' : e20 >= 55 ? 'Healthy' : 'Weakening' },
    { label: '% above 10 EMA',      val: e10.toFixed(1) + '%',  ok: e10 >= 50,   warn: e10 >= 30,  note: e10 < 40 ? 'Pullback risk' : e10 >= 70 ? 'Overbought' : 'Neutral' },
  ];

  var html = '';
  signals.forEach(function(s) {
    var dot = s.ok ? '#22c55e' : (s.warn ? '#f59e0b' : '#ef4444');
    html += '<div class="signal-row">';
    html += '<div class="signal-left">';
    html += '<div class="dot" style="background:' + dot + '"></div>';
    html += '<span class="signal-name">' + s.label + '</span>';
    html += '</div>';
    html += '<div class="signal-right">';
    html += '<span class="signal-val" style="color:' + dot + '">' + s.val + '</span>';
    html += '<span class="signal-note">' + s.note + '</span>';
    html += '</div></div>';
  });
  document.getElementById('signals').innerHTML = html;

  var alert;
  if (divergence) {
    alert = 'Divergence detected — short/medium breadth (' + e20.toFixed(1) + '% / ' + e50.toFixed(1) + '%) is elevated but long term breadth (' + e200.toFixed(1) + '%) has not confirmed. Index may be running on narrow leadership. Watch for failed breakouts and reduce new entries.';
  } else if (stWeak && stLeader) {
    alert = 'Short term breadth weakening (' + e10.toFixed(1) + '%) while Minervini leaders remain strong (' + m9.toFixed(0) + ' stocks). Leaders holding but broader market pulling back. Reduce new entries, protect open positions with tighter stops.';
  } else if (bullConfirmed && stLeader) {
    alert = 'Both Minervini leaders and long term breadth are aligned. This is the ideal environment. Prioritize ELITE setups with full conviction.';
  } else if (!bullConfirmed && e200 >= 40) {
    alert = '% above 200 EMA at ' + e200.toFixed(1) + '% — approaching the 50% bull confirmation threshold. Watch closely. A cross above 50% is a major regime change signal for your system.';
  } else {
    alert = 'Mixed conditions. Let the 200 EMA breadth be your primary guide — stay defensive until it crosses 50%.';
  }
  document.getElementById('alert-text').textContent = alert;
  document.getElementById('output').style.display = 'block';
}

calculate();
</script>
</body>
</html>
""", height=900, scrolling=True)

# ── SCORE GUIDE ───────────────────────────────────────────────
elif st.session_state.page == "score_guide":
    if st.button("← Back to Portal", key="back_sg"):
        st.session_state.page = "home"
        st.rerun()
    st.components.v1.html("""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>SwingEdge Pro — Market Verdict Score Guide</title>
<link href="https://fonts.googleapis.com/css2?family=Instrument+Serif:ital@0;1&family=DM+Mono:wght@400;500&family=Geist:wght@300;400;500&display=swap" rel="stylesheet">
<style>
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
:root {
  --gold: #c9a84c; --gold-light: #e8c97a; --gold-dim: #8a6e2f;
  --navy: #080e1a; --navy-2: #0d1626; --navy-3: #111f35; --navy-4: #1a2d47;
  --text: #dde4f0; --text-dim: #7a8fa8; --text-muted: #3d5470;
  --green: #3dd68c; --amber: #f5a623; --red: #f04e4e; --teal: #2dd4bf;
}
html { scroll-behavior: smooth; }
body { background: var(--navy); color: var(--text); font-family: 'Geist', sans-serif; font-weight: 300; line-height: 1.7; min-height: 100vh; }
.page { max-width: 860px; margin: 0 auto; padding: 3rem 2rem 4rem; }
.top-rule { display: flex; align-items: center; gap: 1rem; margin-bottom: 2.5rem; }
.top-rule::before { content: ''; flex: 1; height: 0.5px; background: linear-gradient(90deg, transparent, var(--gold-dim)); }
.top-rule::after  { content: ''; flex: 1; height: 0.5px; background: linear-gradient(90deg, var(--gold-dim), transparent); }
.top-rule span { font-family: 'DM Mono', monospace; font-size: 10px; letter-spacing: 0.2em; color: var(--gold-dim); text-transform: uppercase; white-space: nowrap; }
.masthead { text-align: center; margin-bottom: 3rem; }
.masthead .brand { font-family: 'DM Mono', monospace; font-size: 10px; letter-spacing: 0.25em; color: var(--gold); text-transform: uppercase; margin-bottom: 1rem; }
.masthead h1 { font-family: 'Instrument Serif', serif; font-size: clamp(2rem, 5vw, 3.2rem); font-weight: 400; color: var(--text); line-height: 1.15; margin-bottom: 0.75rem; }
.masthead h1 em { font-style: italic; color: var(--gold-light); }
.masthead .subtitle { font-size: 14px; color: var(--text-dim); max-width: 520px; margin: 0 auto; }
.score-hero { background: var(--navy-2); border: 0.5px solid var(--navy-4); border-radius: 16px; padding: 2rem; margin-bottom: 2.5rem; position: relative; overflow: hidden; }
.score-hero::before { content: ''; position: absolute; top: 0; left: 0; right: 0; height: 1px; background: linear-gradient(90deg, transparent, var(--gold-dim), transparent); }
.score-hero-label { font-family: 'DM Mono', monospace; font-size: 10px; letter-spacing: 0.18em; color: var(--gold-dim); text-transform: uppercase; margin-bottom: 1.25rem; }
.score-track { position: relative; height: 28px; background: var(--navy-3); border-radius: 4px; overflow: hidden; margin-bottom: 0.5rem; display: flex; }
.score-segment { display: flex; align-items: center; justify-content: center; font-family: 'DM Mono', monospace; font-size: 9px; letter-spacing: 0.06em; }
.score-labels { display: flex; justify-content: space-between; margin-top: 0.5rem; }
.score-labels span { font-family: 'DM Mono', monospace; font-size: 9px; color: var(--text-muted); }
.seg-bear   { width: 25%; background: rgba(240,78,78,0.25);  color: #f04e4e; }
.seg-chop   { width: 15%; background: rgba(245,166,35,0.2);  color: #f5a623; }
.seg-recov  { width: 15%; background: rgba(245,166,35,0.15); color: #f5c94e; }
.seg-bull   { width: 20%; background: rgba(61,214,140,0.15); color: #3dd68c; }
.seg-strong { width: 25%; background: rgba(61,214,140,0.25); color: #2dd4bf; }
.divider { display: flex; align-items: center; gap: 1rem; margin: 2rem 0; }
.divider::before, .divider::after { content: ''; flex: 1; height: 0.5px; background: var(--navy-4); }
.divider span { font-family: 'Instrument Serif', serif; font-size: 11px; font-style: italic; color: var(--text-muted); }
.section-title { font-family: 'DM Mono', monospace; font-size: 10px; letter-spacing: 0.18em; color: var(--gold); text-transform: uppercase; margin-bottom: 1.25rem; }
.level-row { display: grid; grid-template-columns: 80px 150px 1fr 1fr; gap: 0; border-bottom: 0.5px solid var(--navy-4); padding: 0.85rem 0; align-items: start; }
.level-row:first-child { border-top: 0.5px solid var(--navy-4); }
.level-score { font-family: 'DM Mono', monospace; font-size: 13px; font-weight: 500; padding-right: 1rem; }
.level-name  { font-size: 13px; font-weight: 500; padding-right: 1rem; }
.level-meaning { font-size: 12px; color: var(--text-dim); padding-right: 1rem; line-height: 1.5; }
.level-action  { font-size: 12px; color: var(--text-dim); line-height: 1.5; }
.inputs-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 12px; margin-bottom: 2.5rem; }
.input-card { background: var(--navy-2); border: 0.5px solid var(--navy-4); border-radius: 10px; padding: 1rem 1.25rem; }
.input-card-wide { grid-column: 1 / -1; }
.ic-label  { font-family: 'DM Mono', monospace; font-size: 10px; letter-spacing: 0.12em; color: var(--gold-dim); text-transform: uppercase; margin-bottom: 4px; }
.ic-weight { font-size: 22px; font-weight: 500; color: var(--gold-light); line-height: 1; margin-bottom: 4px; }
.ic-source { font-size: 11px; color: var(--text-muted); }
.ic-note   { font-size: 12px; color: var(--text-dim); margin-top: 6px; line-height: 1.4; }
.highlight-box { background: var(--navy-2); border: 0.5px solid var(--gold-dim); border-radius: 12px; padding: 1.5rem; margin-bottom: 2.5rem; position: relative; overflow: hidden; }
.highlight-box::before { content: ''; position: absolute; top: 0; left: 0; right: 0; height: 1px; background: linear-gradient(90deg, transparent, var(--gold), transparent); }
.hb-eyebrow { font-family: 'DM Mono', monospace; font-size: 10px; letter-spacing: 0.18em; color: var(--gold); text-transform: uppercase; margin-bottom: 0.75rem; }
.hb-title { font-family: 'Instrument Serif', serif; font-size: 20px; color: var(--text); margin-bottom: 0.5rem; }
.hb-body  { font-size: 13px; color: var(--text-dim); line-height: 1.7; }
.hb-number { font-family: 'DM Mono', monospace; font-size: 13px; color: var(--gold-light); background: var(--navy-3); display: inline-block; padding: 2px 8px; border-radius: 4px; margin: 0 2px; }
.divergence-box { background: rgba(245,166,35,0.06); border: 0.5px solid rgba(245,166,35,0.25); border-radius: 12px; padding: 1.25rem 1.5rem; margin-bottom: 2.5rem; }
.div-title { font-family: 'DM Mono', monospace; font-size: 10px; letter-spacing: 0.15em; color: var(--amber); text-transform: uppercase; margin-bottom: 0.75rem; }
.div-body  { font-size: 13px; color: var(--text-dim); line-height: 1.7; }
.div-body strong { color: var(--text); font-weight: 500; }
.workflow-step { display: flex; gap: 1.25rem; padding: 0.85rem 0; border-bottom: 0.5px solid var(--navy-4); align-items: flex-start; }
.workflow-step:first-child { border-top: 0.5px solid var(--navy-4); }
.ws-num    { font-family: 'Instrument Serif', serif; font-size: 20px; font-style: italic; color: var(--gold-dim); min-width: 28px; line-height: 1; padding-top: 2px; }
.ws-title  { font-size: 13px; font-weight: 500; color: var(--text); margin-bottom: 2px; }
.ws-detail { font-size: 12px; color: var(--text-dim); line-height: 1.5; }
.rule { display: flex; gap: 1rem; padding: 0.85rem 0; border-bottom: 0.5px solid var(--navy-4); align-items: flex-start; }
.rule:first-child { border-top: 0.5px solid var(--navy-4); }
.rule-num  { font-family: 'DM Mono', monospace; font-size: 11px; color: var(--gold-dim); min-width: 24px; padding-top: 2px; }
.rule-text { font-size: 13px; color: var(--text-dim); line-height: 1.6; }
.rule-text strong { color: var(--text); font-weight: 500; }
.limit-item { display: flex; gap: 0.75rem; padding: 0.6rem 0; font-size: 13px; color: var(--text-dim); border-bottom: 0.5px solid var(--navy-4); line-height: 1.5; align-items: flex-start; }
.limit-item:first-child { border-top: 0.5px solid var(--navy-4); }
.limit-x { font-family: 'DM Mono', monospace; font-size: 11px; color: var(--red); padding-top: 2px; flex-shrink: 0; }
.tagline { text-align: center; padding: 2rem 0 0; border-top: 0.5px solid var(--navy-4); }
.tagline p { font-family: 'Instrument Serif', serif; font-size: 15px; font-style: italic; color: var(--text-muted); }
.tagline strong { font-style: normal; color: var(--gold); font-weight: 400; }
.mb-25 { margin-bottom: 2.5rem; }
</style>
</head>
<body>
<div class="page">

<div class="top-rule"><span>Knowledge Pager</span></div>

<div class="masthead">
  <div class="brand">SwingEdge Pro &mdash; Market Intelligence</div>
  <h1>The Market Verdict<br><em>Score Explained</em></h1>
  <p class="subtitle">A composite breadth score that tells you how much risk to take. Not what the market will do, but how confidently to act right now.</p>
</div>

<div class="score-hero">
  <div class="score-hero-label">Composite score range &mdash; 0 to 100</div>
  <div class="score-track">
    <div class="score-segment seg-bear">Bear</div>
    <div class="score-segment seg-chop">Choppy</div>
    <div class="score-segment seg-recov">Recov</div>
    <div class="score-segment seg-bull">Bull</div>
    <div class="score-segment seg-strong">Strong Bull</div>
  </div>
  <div class="score-labels">
    <span>0</span><span>25</span><span>40</span><span>55</span><span>65</span><span>80</span><span>100</span>
  </div>
</div>

<div class="section-title">Score levels and what they mean</div>
<div class="mb-25">
  <div class="level-row">
    <div class="level-score" style="color:#f04e4e">0 to 25</div>
    <div class="level-name" style="color:#f04e4e">Bear / avoid</div>
    <div class="level-meaning">Market structure broken. Most stocks in downtrend. Rallies are traps.</div>
    <div class="level-action">100% cash. No longs under any circumstance.</div>
  </div>
  <div class="level-row">
    <div class="level-score" style="color:#f5a623">25 to 40</div>
    <div class="level-name" style="color:#f5a623">Weak / correction</div>
    <div class="level-meaning">Breadth deteriorating. Leaders struggling to hold breakouts.</div>
    <div class="level-action">No new trades. Tighten stops on all open positions.</div>
  </div>
  <div class="level-row">
    <div class="level-score" style="color:#f5c94e">40 to 55</div>
    <div class="level-name" style="color:#f5c94e">Choppy / recovering</div>
    <div class="level-meaning">Mixed signals. Some leaders working, most stocks going nowhere.</div>
    <div class="level-action">ELITE only. 25 to 50% size. Exit fast on failures.</div>
  </div>
  <div class="level-row">
    <div class="level-score" style="color:#3dd68c">55 to 65</div>
    <div class="level-name" style="color:#3dd68c">Selective bull</div>
    <div class="level-meaning">Strong leadership present. Broad market not yet fully confirmed.</div>
    <div class="level-action">ELITE and PRIME. Half to 75% size. Watch 200 EMA breadth.</div>
  </div>
  <div class="level-row">
    <div class="level-score" style="color:#3dd68c">65 to 80</div>
    <div class="level-name" style="color:#3dd68c">Healthy bull</div>
    <div class="level-meaning">Leaders strong. Breadth improving. Most breakouts sustaining.</div>
    <div class="level-action">ELITE, PRIME, STRONG. Normal sizing 75 to 100%.</div>
  </div>
  <div class="level-row">
    <div class="level-score" style="color:#2dd4bf">80 to 100</div>
    <div class="level-name" style="color:#2dd4bf">Full bull</div>
    <div class="level-meaning">Broad participation confirmed. Even average setups working.</div>
    <div class="level-action">All setups. Full positions. Add to winners aggressively.</div>
  </div>
</div>

<div class="divider"><span>the five inputs</span></div>

<div class="section-title">How the score is built</div>
<div class="inputs-grid">
  <div class="input-card input-card-wide">
    <div class="ic-label">9/9 Minervini count &mdash; 35 points</div>
    <div class="ic-weight">35%</div>
    <div class="ic-source">Source: SwingEdge Python scanner (automatic after 3:30 PM)</div>
    <div class="ic-note">Stocks passing all 9 Minervini SEPA conditions. Measures leadership quality, not quantity. Capped at 300 for scoring. Above 150 means strong leaders are present. Above 200 is super bull leadership territory.</div>
  </div>
  <div class="input-card">
    <div class="ic-label">% above 200 EMA &mdash; 30 points</div>
    <div class="ic-weight">30%</div>
    <div class="ic-source">Source: SwingEdge scanner or Chartink table</div>
    <div class="ic-note">The most important single input. Crossing 50% is bull market confirmation. This one number changes the entire regime verdict.</div>
  </div>
  <div class="input-card">
    <div class="ic-label">% above 50 EMA &mdash; 20 points</div>
    <div class="ic-weight">20%</div>
    <div class="ic-source">Source: SwingEdge scanner or Chartink table</div>
    <div class="ic-note">Medium term participation. Above 60% is healthy. Above 80% is overbought and a pause or pullback is likely soon.</div>
  </div>
  <div class="input-card">
    <div class="ic-label">% above 20 EMA &mdash; 10 points</div>
    <div class="ic-weight">10%</div>
    <div class="ic-source">Source: Chartink Abv 20ma column</div>
    <div class="ic-note">Short term momentum. Fast-moving. Dropping below 40% from high levels signals near-term weakness even in a bull market.</div>
  </div>
  <div class="input-card">
    <div class="ic-label">% above 10 EMA &mdash; 5 points</div>
    <div class="ic-weight">5%</div>
    <div class="ic-source">Source: Chartink Abv 10ma column</div>
    <div class="ic-note">Shortest term pulse. Drops and recovers quickly. Sharp fall signals a 3 to 7 day shakeout coming. Do not panic sell on this alone.</div>
  </div>
</div>

<div class="divider"><span>the number that matters most</span></div>

<div class="highlight-box">
  <div class="hb-eyebrow">Critical threshold</div>
  <div class="hb-title">% stocks above 200 EMA crossing 50%</div>
  <div class="hb-body">
    This is the single most important signal in the entire system. When less than half the market is above its 200 EMA, you are in a <span class="hb-number">recovery or correction</span> regime regardless of what the index is doing. When it crosses <span class="hb-number">50%</span>, the bull market is officially confirmed. Increase position sizes, widen stops, let winners run. Watch this number daily. Right now at <span class="hb-number">46.5%</span> you are 3.5 points from a regime change.
  </div>
</div>

<div class="divergence-box">
  <div class="div-title">Divergence alert &mdash; the hidden warning signal</div>
  <div class="div-body">A divergence occurs when <strong>short or medium term breadth is elevated above 70%</strong> but <strong>long term breadth is still below 50%</strong>. This means the index is being carried by a concentrated group of heavy stocks while most stocks quietly weaken underneath. Action when divergence is detected: <strong>do not add new positions, tighten stops on all open trades, reduce immediately on any failed breakout.</strong> This exact pattern preceded the Feb to Apr 2026 correction.</div>
</div>

<div class="section-title">Daily workflow &mdash; 5 minutes after 3:30 PM</div>
<div class="mb-25">
  <div class="workflow-step">
    <div class="ws-num">1</div>
    <div>
      <div class="ws-title">Run your Python scanner</div>
      <div class="ws-detail">market_breadth_engine.py calculates % above 200, 50, 20 EMA and the 9/9 Minervini count automatically from your 1,310 stock universe.</div>
    </div>
  </div>
  <div class="workflow-step">
    <div class="ws-num">2</div>
    <div>
      <div class="ws-title">Check Chartink for % above 10 EMA</div>
      <div class="ws-detail">Your breadth table at chartink.com/dashboard/356290 &mdash; read today's Abv 10ma figure from the top row.</div>
    </div>
  </div>
  <div class="workflow-step">
    <div class="ws-num">3</div>
    <div>
      <div class="ws-title">Enter all 5 numbers into the verdict tool</div>
      <div class="ws-detail">Open SwingEdge_Market_Verdict.html in your browser. Enter the 5 numbers. Hit calculate. Read the verdict in under 30 seconds.</div>
    </div>
  </div>
  <div class="workflow-step">
    <div class="ws-num">4</div>
    <div>
      <div class="ws-title">Apply the sizing rule to every trade that day</div>
      <div class="ws-detail">Not selectively. Every new position follows the rule without exception. The score only works if you apply it consistently.</div>
    </div>
  </div>
  <div class="workflow-step">
    <div class="ws-num">5</div>
    <div>
      <div class="ws-title">Share the verdict with your members</div>
      <div class="ws-detail">Post the score, verdict name, and sizing rule in the SwingEdge community. One line. Every day. Builds trust and keeps members disciplined in all market conditions.</div>
    </div>
  </div>
</div>

<div class="section-title">Non-negotiable rules</div>
<div class="mb-25">
  <div class="rule"><div class="rule-num">01</div><div class="rule-text"><strong>Score below 40 means zero new longs.</strong> No exceptions. Not even your highest conviction ELITE setup. The environment is not rewarding risk right now. Wait.</div></div>
  <div class="rule"><div class="rule-num">02</div><div class="rule-text"><strong>200 EMA breadth is the regime arbiter.</strong> Until it crosses 50%, you are in selective bull at best. A high Minervini count alone cannot override this. Leaders can hold while the broad market breaks down beneath them.</div></div>
  <div class="rule"><div class="rule-num">03</div><div class="rule-text"><strong>Score dropping 10 or more points in 3 days means reduce exposure immediately.</strong> Do not wait for stabilization. Cut size on all new entries and tighten stops on existing positions.</div></div>
  <div class="rule"><div class="rule-num">04</div><div class="rule-text"><strong>When 10 EMA breadth drops below 40%, expect a 3 to 7 day shakeout.</strong> Even in a healthy bull market. Use the dip to find better entries, not to panic sell strong positions.</div></div>
  <div class="rule"><div class="rule-num">05</div><div class="rule-text"><strong>The score is a risk filter, not a buy signal.</strong> A score of 75 does not mean buy anything. It means the environment supports full-sized positions in your best quality setups only.</div></div>
</div>

<div class="section-title">What this score cannot do</div>
<div class="mb-25">
  <div class="limit-item"><span class="limit-x">x</span>Predict market direction or how long any regime will last.</div>
  <div class="limit-item"><span class="limit-x">x</span>Tell you which specific stocks to buy or exactly when to enter them.</div>
  <div class="limit-item"><span class="limit-x">x</span>Prevent drawdowns &mdash; even a score of 80 does not guarantee breakouts work.</div>
  <div class="limit-item"><span class="limit-x">x</span>Replace your chart reading, RS analysis, or setup quality judgment.</div>
  <div class="limit-item"><span class="limit-x">x</span>Account for sudden macro shocks, geopolitical events, or circuit breakers.</div>
</div>

<div class="tagline">
  <p><strong>SwingEdge Pro</strong> &mdash; Your Edge. Every Week.</p>
</div>

</div>
</body>
</html>
""", height=900, scrolling=True)

# ── BOTTOM BOUNCE DASHBOARD ───────────────────────────────────
elif st.session_state.page == "bottom_bounce":

    @st.cache_data(ttl=300)
    def load_bb():
        try:
            api = requests.get("https://api.github.com/repos/SwingEdgeLab/swingEdge-dashboard/contents/").json()
            bb_files = sorted([f['name'] for f in api if isinstance(f, dict) and f['name'].startswith('SwingEdge_BottomBounce') and f['name'].endswith('.xlsx')], reverse=True)
            filename = bb_files[0] if bb_files else 'bottom_bounce.xlsx'
            url = f"https://github.com/SwingEdgeLab/swingEdge-dashboard/raw/main/{filename}"
            r = requests.get(url, headers={"Accept": "application/octet-stream"})
            wb = openpyxl.load_workbook(BytesIO(r.content), data_only=True)

            ws_info = wb['📅 Scan Info']
            info = {}
            for row in ws_info.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    info[str(row[0])] = str(row[1])

            sheet_map = {
                'T0': '👀 T0 — Pre-Recovery',
                'T1': '🌱 T1 — 10W Cross 20W',
                'T2': '🔥 T2 — 10W Cross 40W',
                'T3': '🚀 T3 — 20W Cross 40W',
                'COMBO': '⚡ COMBO T2+T3',
            }
            data = {}
            for key, sheetname in sheet_map.items():
                if sheetname in wb.sheetnames:
                    ws = wb[sheetname]
                    rows = list(ws.iter_rows(values_only=True))
                    if len(rows) < 2:
                        data[key] = pd.DataFrame()
                        continue
                    headers = [str(h) if h is not None else f'_c{i}' for i, h in enumerate(rows[0])]
                    df = pd.DataFrame(rows[1:], columns=headers)
                    if headers[0].startswith('_c'):
                        df = df.iloc[:, 1:]
                    df = df[df['Symbol'].notna()]
                    data[key] = df
                else:
                    data[key] = pd.DataFrame()
            return data, info
        except Exception as e:
            return None, None

    if st.button("← Back to Portal", key="back"):
        st.session_state.page = "home"
        st.rerun()

    with st.spinner("Loading scan data..."):
        data, info = load_bb()

    if data is None:
        st.error("Could not load scan data. Please check GitHub repo.")
        st.stop()

    scan_date = info.get('Scan Date', '—')
    run_time = info.get('Run Time', '—')
    counts = {k: len(data.get(k, pd.DataFrame())) for k in ['T0','T1','T2','T3','COMBO']}
    total = sum(counts[k] for k in ['T0','T1','T2','T3'])

    st.markdown(f"""
    <div style="padding: 4.5rem 1rem 0.5rem; display:flex; align-items:center; justify-content:space-between; flex-wrap:wrap; gap:0.5rem;">
      <div>
        <span style="font-family:'Barlow Condensed',sans-serif; font-size:1.1rem; font-weight:800; letter-spacing:2px; text-transform:uppercase; color:#f0f0f0;">📊 Bottom <span style='color:#f5a623;'>Bounce</span></span>
        <span style="font-size:0.72rem; color:#666; margin-left:1rem; font-family:'Barlow Condensed',sans-serif; letter-spacing:1px;">{scan_date} · {run_time} IST · 1304 stocks</span>
      </div>
      <div style="display:flex; gap:1.5rem; font-family:'Barlow Condensed',sans-serif; font-size:0.72rem; letter-spacing:1px;">
        <span style="color:#888;">T0 <b style="color:#f5a623;">{counts['T0']}</b></span>
        <span style="color:#888;">T1 <b style="color:#f5a623;">{counts['T1']}</b></span>
        <span style="color:#888;">T2 <b style="color:#f5a623;">{counts['T2']}</b></span>
        <span style="color:#888;">T3 <b style="color:#f5a623;">{counts['T3']}</b></span>
        <span style="color:#888;">COMBO <b style="color:#fbbf24;">{counts['COMBO']}</b></span>
        <span style="color:#888;">TOTAL <b style="color:#f5a623;">{total}</b></span>
      </div>
    </div>
    """, unsafe_allow_html=True)

    stage_keys = ['T0','T1','T2','T3','COMBO']
    desc_map = {
        'T0': 'Pre-Recovery — Gap narrowing, EMAs still inverted',
        'T1': 'Early Cross — 10W has crossed above 20W',
        'T2': 'Momentum — 10W has crossed above 40W',
        'T3': 'Trend Restored — 20W has crossed above 40W',
        'COMBO': 'High Conviction — T2 + T3 both fired this week',
    }

    with st.container():
        st.markdown('<div style="padding: 0 2.5rem;">', unsafe_allow_html=True)
        tabs = st.tabs([
            f"T0  ({counts['T0']})",
            f"T1  ({counts['T1']})",
            f"T2  ({counts['T2']})",
            f"T3  ({counts['T3']})",
            f"COMBO  ({counts['COMBO']})",
        ])
        for tab, key in zip(tabs, stage_keys):
            with tab:
                df = data.get(key, pd.DataFrame())
                if df.empty:
                    st.info("No signals for this stage today.")
                else:
                    st.caption(f"_{desc_map[key]}_")
                    c1, c2 = st.columns([4, 1])
                    with c1:
                        search = st.text_input("", key=f"s_{key}", placeholder="Search symbol...")
                    if search:
                        df = df[df['Symbol'].astype(str).str.contains(search.upper(), na=False)]
                    st.dataframe(df, height=min(900, max(400, len(df)*38)), hide_index=True, use_container_width=True)
                    buf = BytesIO()
                    df.to_excel(buf, index=False)
                    with c2:
                        st.download_button(f"Export {key}", data=buf.getvalue(),
                            file_name=f"BB_{key}_{scan_date}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"dl_{key}")
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="footer">SwingEdgePro.in · Confidential · Member Use Only · © 2026</div>', unsafe_allow_html=True)

# ── ACCUMULATION ──────────────────────────────────────────────
elif st.session_state.page == "accumulation":
    render_generic_scanner(
        page_key     = "accumulation",
        title        = "Accumulation",
        icon         = "🏦",
        accent_color = "#f5a623",
        file_prefix  = "Accumulation_Scanner_",
        sheet_config = [
            ("🎯 Prime Setups",  "Prime Setups — Near ATH + high score"),
            ("🔥 Elite Accum",   "Elite Accum — Highest conviction"),
            ("💪 Strong Accum",  "Strong Accum — Solid structure"),
            ("🌀 Coiling",       "Coiling — Tightening range"),
            ("📈 Above EMA200",  "Above EMA200 — Trend confirmed"),
            ("📋 All Qualified", "All Qualified — Full list"),
        ],
        desc_map = {
            "🎯 Prime Setups":  "Near ATH + accumulation score ≥ threshold — highest conviction setups",
            "🔥 Elite Accum":   "Top-tier accumulation: up days dominating, 1.5x+ volume on buy days",
            "💪 Strong Accum":  "Strong institutional footprint — consistent volume on up days",
            "🌀 Coiling":       "Range contraction with accumulation pattern — pre-breakout tightening",
            "📈 Above EMA200":  "All accumulation signals trading above 200 EMA — trend-confirmed",
            "📋 All Qualified": "All stocks passing base accumulation criteria",
        }
    )

# ── ATH SCANNER ───────────────────────────────────────────────
elif st.session_state.page == "ath_scanner":
    render_generic_scanner(
        page_key     = "ath_scanner",
        title        = "ATH Scanner",
        icon         = "⚡",
        accent_color = "#ffd066",
        file_prefix  = "ATH_Scanner_v7_",
        sheet_config = [
            ("🏅 Combo 5%",       "Combo 5% — Within 5% ATH + quality gates"),
            ("🏅 Combo 10%",      "Combo 10% — Within 10% ATH + quality gates"),
            ("🏅 Combo 20%",      "Combo 20% — Within 20% ATH + quality gates"),
            ("🔥 Within 5% ATH",  "Within 5% ATH — Nearest to all-time high"),
            ("💪 Within 10% ATH", "Within 10% ATH"),
            ("✅ Within 20% ATH", "Within 20% ATH"),
            ("🏆 ATH + Pattern",  "ATH + Pattern — Best technical setups"),
        ],
        desc_map = {
            "🏅 Combo 5%":       "Within 5% of ATH + ADR≥3% + Liquidity≥₹5Cr + EMA50/200 aligned",
            "🏅 Combo 10%":      "Within 10% of ATH + ADR≥3% + Liquidity≥₹5Cr + EMA50/200 aligned",
            "🏅 Combo 20%":      "Within 20% of ATH + ADR≥3% + Liquidity≥₹5Cr + EMA50/200 aligned",
            "🔥 Within 5% ATH":  "Stocks within 5% of their all-time high (ADR≥3% + Liq≥₹5Cr applied)",
            "💪 Within 10% ATH": "Stocks within 10% of their all-time high",
            "✅ Within 20% ATH": "Stocks within 20% of their all-time high",
            "🏆 ATH + Pattern":  "Near-ATH stocks with confirmed patterns (VCP / Inside Bar / 3WT)",
        }
    )

# ── 200 EMA CROSS ─────────────────────────────────────────────
elif st.session_state.page == "ema200_cross":
    render_generic_scanner(
        page_key     = "ema200_cross",
        title        = "200 EMA Cross",
        icon         = "📈",
        accent_color = "#22c55e",
        file_prefix  = "EMA200_Cross_",
        sheet_config = [
            ("🔥 Fresh (1-",   "Fresh — Crossed 1-5 days ago"),
            ("✅ Recent (6-",  "Recent — Crossed 6-20 days ago"),
            ("📋 All Crosses", "All Crosses — Complete universe"),
        ],
        desc_map = {
            "🔥 Fresh (1-":   "Price crossed above 200 EMA within last 5 days — volume confirmed + RS≥75",
            "✅ Recent (6-":  "Crosses 6-20 days old — still actionable on pullbacks",
            "📋 All Crosses": "All 200 EMA crossover signals in the universe",
        }
    )

# ── DEEP BASE RECOVERY ────────────────────────────────────────
elif st.session_state.page == "deep_base":
    render_generic_scanner(
        page_key     = "deep_base",
        title        = "Deep Base Recovery",
        icon         = "🔭",
        accent_color = "#60a5fa",
        file_prefix  = "Deep_Base_Recovery_v1_",
        sheet_config = [
            ("🟢 Very Close",   "Very Close — Within 10% of breakout"),
            ("🔵 Approaching",  "Approaching — Within 25% of breakout"),
            ("📊 All Deep Base","All — Complete scan results"),
        ],
        desc_map = {
            "🟢 Very Close":   "Stage 1 stocks within 10% of prior base high — imminent breakout candidates",
            "🔵 Approaching":  "Within 25% of base high — building momentum, early watch",
            "📊 All Deep Base":"All stocks recovering from extended Stage 1 bases (26+ weeks consolidation)",
        }
    )

# ── 52W HIGH BREAKOUT ─────────────────────────────────────────
elif st.session_state.page == "52w_high":
    render_generic_scanner(
        page_key     = "52w_high",
        title        = "52W High Breakout",
        icon         = "🏔️",
        accent_color = "#a78bfa",
        file_prefix  = "52W_High_Scanner_v1_",
        sheet_config = [
            ("🚀 Breaking Now",      "Breaking Now — At 52W high today"),
            ("🎯 Approaching ≤5%",   "Approaching — Within 5% of 52W high"),
            ("💥 RS Surge RS≥75",    "RS Surge — RS≥75 with momentum"),
            ("🏆 Freshout 26W Base", "Freshout — Breaking out of 26W+ base"),
            ("🔄 ATH Recovery",       "ATH Recovery — Recovering to prior highs"),
        ],
        desc_map = {
            "🚀 Breaking Now":      "Currently printing at or above 52-week high with volume expansion",
            "🎯 Approaching ≤5%":   "Within 5% of 52-week high — potential breakout setups",
            "💥 RS Surge RS≥75":    "RS≥75 with price approaching 52-week highs",
            "🏆 Freshout 26W Base": "Breaking out of a 26+ week base — Weinstein Stage 2 entry",
            "🔄 ATH Recovery":       "Recovering back toward all-time highs from prior correction",
        }
    )

# ── 50 EMA SHAKEOUT ───────────────────────────────────────────
elif st.session_state.page == "50ema_shakeout":
    render_generic_scanner(
        page_key     = "50ema_shakeout",
        title        = "50 EMA Shakeout",
        icon         = "💥",
        accent_color = "#fb923c",
        file_prefix  = "50EMA_Shakeout_v1_",
        sheet_config = [
            ("🔥 Reclaimed Today",    "Reclaimed Today — Bounced off 50 EMA today"),
            ("🎯 Reclaimed 1-3 Days", "Reclaimed 1-3 Days — Recent bounce"),
        ],
        desc_map = {
            "🔥 Reclaimed Today":    "Stage 2 leaders that touched/undercut 50 EMA and reclaimed it today",
            "🎯 Reclaimed 1-3 Days": "Reclaimed 50 EMA 1-3 days ago — entry window still open",
        }
    )

# ── RS EXPLOSION ──────────────────────────────────────────────
elif st.session_state.page == "rs_explosion":
    render_generic_scanner(
        page_key     = "rs_explosion",
        title        = "RS Explosion",
        icon         = "🚀",
        accent_color = "#f5a623",
        file_prefix  = "RS_Explosion_",
        sheet_config = [
            ("🔥 Explosive",     "Explosive — Massive RS surge"),
            ("⚡ Strong",        "Strong — Significant RS jump"),
            ("💎 Combo RS+ADR",  "Combo — RS surge + ADR≥3%"),
            ("🚀 EMA200 Crossed","EMA200 Crossed — RS surge + above 200 EMA"),
            ("📈 Valid",         "Valid — All qualifying RS moves"),
            ("📊 All Results",   "All Results — Complete universe"),
        ],
        desc_map = {
            "🔥 Explosive":      "RS jumped explosively — institutions aggressively accumulating",
            "⚡ Strong":         "Strong RS surge — meaningful relative strength improvement",
            "💎 Combo RS+ADR":   "RS surge with ADR≥3% — volatile enough to trade profitably",
            "🚀 EMA200 Crossed": "RS surge AND price just crossed above 200 EMA — regime change",
            "📈 Valid":          "All valid RS explosion signals passing minimum thresholds",
            "📊 All Results":    "Complete ranked output for the full universe",
        }
    )

# ── L1 STRUCTURAL QUALITY ─────────────────────────────────────
elif st.session_state.page == "l1_structural":
    render_generic_scanner(
        page_key     = "l1_structural",
        title        = "L1 Structural Quality",
        icon         = "🏆",
        accent_color = "#ffd066",
        file_prefix  = "L1_Structural_Quality_",
        sheet_config = [
            ("🏆 ELITE",          "ELITE — 9/9 Minervini conditions"),
            ("⚡ PRIME",          "PRIME — 7-8/9 conditions"),
            ("✅ STRONG",         "STRONG — 5-6/9 conditions"),
            ("👁 WATCH",          "WATCH — On radar"),
            ("⭐ COMBO",          "COMBO — All grades above EMA200"),
            ("🔥 VPA Confluence", "VPA Confluence — Volume-Price signals"),
            ("🚀 Momentum Leaders","Momentum Leaders — Percentile ranked MRS"),
            ("📊 Sector Strength","Sector Strength — Heatmap"),
            ("📋 ALL",            "ALL — Complete universe"),
        ],
        desc_map = {
            "🏆 ELITE":           "All 9 Minervini SEPA conditions met — highest conviction watchlist",
            "⚡ PRIME":           "7-8 of 9 conditions met — strong structure, near breakout",
            "✅ STRONG":          "5-6 of 9 conditions — solid momentum, worth monitoring",
            "👁 WATCH":           "On radar — improving structure, not yet qualifying",
            "⭐ COMBO":           "All grades where price is above EMA200 — trend confirmed",
            "🔥 VPA Confluence":  "Volume-Price Analysis signals — U/D ratio + RVol + VDU",
            "🚀 Momentum Leaders":"Percentile-ranked MRS composite score — top relative strength leaders",
            "📊 Sector Strength": "Sector heatmap showing money flow and sector momentum",
            "📋 ALL":             "Full graded universe — all stocks with scores",
        }
    )

# ── L2 INSTITUTIONAL VOLUME ───────────────────────────────────
elif st.session_state.page == "l2_institutional":
    render_generic_scanner(
        page_key     = "l2_institutional",
        title        = "L2 Institutional Volume",
        icon         = "🏛️",
        accent_color = "#22c55e",
        file_prefix  = "L2_Institutional_Volume_",
        sheet_config = [
            ("🔥 CLIMAX",  "CLIMAX — Composite≥85 + heavy volume event"),
            ("🏆 TIER 1",  "TIER 1 — Composite≥70 + high quality volume"),
            ("⚡ TIER 2",  "TIER 2 — Composite≥55"),
            ("👁 WATCH",   "WATCH — Composite≥40"),
            ("📊 Vol Events","Vol Events — All volume events"),
            ("📋 ALL",     "ALL — Complete results"),
        ],
        desc_map = {
            "🔥 CLIMAX":    "Highest conviction — composite≥85, HVE/HVY volume fired, VCP structure confirmed",
            "🏆 TIER 1":    "High quality institutional accumulation — composite≥70 with HVQ event",
            "⚡ TIER 2":    "Solid institutional interest — composite≥55",
            "👁 WATCH":     "Early accumulation signal — composite≥40, monitoring stage",
            "📊 Vol Events":"All volume events detected in the universe",
            "📋 ALL":       "Complete graded results across all stocks",
        }
    )

# ── RS SCREENER ───────────────────────────────────────────────
elif st.session_state.page == "rs_screener":
    render_generic_scanner(
        page_key     = "rs_screener",
        title        = "RS Screener",
        icon         = "📡",
        accent_color = "#a78bfa",
        file_prefix  = "RS_Screener_v2_",
        sheet_config = [
            ("🔥 RS90+ ADR3.5+",       "RS90+ ADR3.5+ — Elite RS with volatility"),
            ("🎯 Structure Hunt 65-90", "Structure Hunt — RS 65-90 building bases"),
            ("✅ RS+All EMAs (BEST)",   "RS + All EMAs — Best quality setups"),
            ("All Stocks Ranked",       "All Stocks Ranked — Full universe by RS"),
        ],
        desc_map = {
            "🔥 RS90+ ADR3.5+":        "RS≥90 + ADR≥3.5% — elite relative strength with enough volatility to trade",
            "🎯 Structure Hunt 65-90":  "RS 65-90 range — strong but not yet elite, building bases for future breakout",
            "✅ RS+All EMAs (BEST)":    "RS qualified + price above all 4 EMAs — trend fully confirmed",
            "All Stocks Ranked":        "Full universe ranked by RS score — spot emerging leaders early",
        }
    )

# ── EP SCANNER ────────────────────────────────────────────────
elif st.session_state.page == "ep_scanner":
    render_generic_scanner(
        page_key     = "ep_scanner",
        title        = "EP Scanner",
        icon         = "⚡",
        accent_color = "#f87171",
        file_prefix  = "EP_Scanner_",
        sheet_config = [
            ("🔥 EXPLOSIVE", "EXPLOSIVE — Volume≥10x average"),
            ("⚡ STRONG EP", "STRONG EP — Volume≥7x average"),
            ("✅ VALID EP",  "VALID EP — Volume≥3x average"),
            ("📋 ALL EPs",   "ALL EPs — Every qualifying pivot"),
        ],
        desc_map = {
            "🔥 EXPLOSIVE": "Volume ≥10x average — rare, highest urgency episodic pivot",
            "⚡ STRONG EP":  "Volume ≥7x average — strong institutional conviction pivot",
            "✅ VALID EP":   "Volume ≥3x average — valid pivot, worth monitoring for follow-through",
            "📋 ALL EPs":    "All episodic pivot candidates — complete list",
        }
    )

# ── HTF SCANNER ───────────────────────────────────────────────
elif st.session_state.page == "htf_scanner":
    render_generic_scanner(
        page_key     = "htf_scanner",
        title        = "HTF Scanner",
        icon         = "🎯",
        accent_color = "#34d399",
        file_prefix  = "HTF_Scanner_RELAXED_",
        sheet_config = [
            ("HTF Patterns", "HTF Patterns — Confirmed High Tight Flag setups"),
            ("All Stocks",   "All Stocks — Full scan universe"),
        ],
        desc_map = {
            "HTF Patterns": "Confirmed High Tight Flag — parabolic move followed by tight ≤25% consolidation",
            "All Stocks":   "Full universe scan results with HTF scoring",
        }
    )

# ── 3WTC / NR SCANNER ─────────────────────────────────────────
elif st.session_state.page == "nr_3wtc":
    render_generic_scanner(
        page_key     = "nr_3wtc",
        title        = "3WTC / NR Scanner",
        icon         = "📐",
        accent_color = "#fb923c",
        file_prefix  = "3WTC_NR_Scanner_v1_2_",
        sheet_config = [
            ("3WTC Prime", "3WTC Prime — Tightest closes, highest conviction"),
            ("3WTC All",   "3WTC All — All 3-Week Tight Close patterns"),
            ("NR4",        "NR4 — Narrowest range in 4 days"),
            ("NR7",        "NR7 — Narrowest range in 7 days"),
        ],
        desc_map = {
            "3WTC Prime": "3 consecutive weeks closing within 1.5% of each other — institutional coiling before breakout",
            "3WTC All":   "All 3-Week Tight Close patterns — price coiling tightly over 3 weeks",
            "NR4":        "Narrowest daily range in last 4 days — short-term volatility contraction",
            "NR7":        "Narrowest daily range in last 7 days — stronger volatility contraction signal",
        }
    )

# ── WEEKLY SCANNER ────────────────────────────────────────────
elif st.session_state.page == "weekly_scanner":
    render_generic_scanner(
        page_key     = "weekly_scanner",
        title        = "Weekly Scanner",
        icon         = "📅",
        accent_color = "#60a5fa",
        file_prefix  = "Weekly_Scanner_v3_",
        sheet_config = [
            ("🏆 Power Breakouts",  "Power Breakouts — Strongest weekly signals"),
            ("✅ Valid Breakouts",  "Valid Breakouts — All qualifying breakouts"),
            ("🔒 3WTC",            "3WTC — 3-Week Tight Close"),
            ("🔲 Inside Week",     "Inside Week — Weekly inside bar"),
            ("🌀 Mini Coil Daily", "Mini Coil Daily — Daily coiling"),
            ("🌀 Mini Coil Weekly","Mini Coil Weekly — Weekly coiling"),
            ("📐 NR7 Tightening",  "NR7 Tightening — Narrowest range in 7 days"),
        ],
        desc_map = {
            "🏆 Power Breakouts":  "Highest conviction weekly breakouts — volume confirmed, RS strong",
            "✅ Valid Breakouts":  "All qualifying weekly breakout setups",
            "🔒 3WTC":            "3 consecutive weeks closing within tight range — pre-breakout coil",
            "🔲 Inside Week":     "This week's range fully inside last week — volatility compression",
            "🌀 Mini Coil Daily": "Daily mini coil pattern — tightening before next move",
            "🌀 Mini Coil Weekly":"Weekly mini coil — larger timeframe compression",
            "📐 NR7 Tightening":  "Narrowest 7-day range — volatility at multi-week lows",
        }
    )

# ── ULTRAPRO WEEKLY ───────────────────────────────────────────
elif st.session_state.page == "ultrapro_weekly":
    render_generic_scanner(
        page_key     = "ultrapro_weekly",
        title        = "UltraPro Weekly",
        icon         = "💎",
        accent_color = "#ffd066",
        file_prefix  = "UltraPro_Weekly_v4_2_",
        sheet_config = [
            ("ULTRA PRIME", "ULTRA PRIME — Highest conviction weekly setups"),
            ("STRONG",      "STRONG — Strong weekly momentum"),
            ("WATCH",       "WATCH — On radar"),
            ("All Signals", "All Signals — Complete results"),
        ],
        desc_map = {
            "ULTRA PRIME": "Top-tier weekly momentum — all conditions met with volume confirmation",
            "STRONG":      "Strong weekly setup — most conditions met",
            "WATCH":       "Building toward a setup — monitor closely",
            "All Signals": "Complete UltraPro Weekly scan output",
        }
    )

# ── VDU STAGE 2 ───────────────────────────────────────────────
elif st.session_state.page == "vdu_stage2":
    render_generic_scanner(
        page_key     = "vdu_stage2",
        title        = "VDU Stage 2",
        icon         = "📶",
        accent_color = "#34d399",
        file_prefix  = "VDU_Stage2_Scanner_v2_",
        sheet_config = [
            ("🥇 Elite Setups",      "Elite — Highest VDU conviction"),
            ("🔒 3WTC & 3MTC",       "3WTC & 3MTC — Tight close patterns"),
            ("📈 Stage2 Established","Stage 2 Established — Confirmed uptrend"),
            ("🚀 Stage2 Transition", "Stage 2 Transition — Just entering Stage 2"),
            ("📋 All Qualified",     "All Qualified — Full list"),
            ("📊 Score Ranked",      "Score Ranked — By VDU composite"),
        ],
        desc_map = {
            "🥇 Elite Setups":       "Top VDU score — strong volume demand urgency in Stage 2 structure",
            "🔒 3WTC & 3MTC":        "3-Week/Month Tight Close within Stage 2 — highest-probability entry",
            "📈 Stage2 Established": "Firmly in Stage 2 uptrend — EMA stack aligned, RS strong",
            "🚀 Stage2 Transition":  "Just crossed into Stage 2 — early entry opportunity",
            "📋 All Qualified":      "All stocks passing VDU Stage 2 minimum criteria",
            "📊 Score Ranked":       "Full universe ranked by VDU composite score",
        }
    )

# ── RESILIENT STACK REVERSAL ──────────────────────────────────
elif st.session_state.page == "resilient_stack":
    render_generic_scanner(
        page_key     = "resilient_stack",
        title        = "Resilient Stack Reversal",
        icon         = "🛡️",
        accent_color = "#a78bfa",
        file_prefix  = "Resilient_Stack_Reversal_",
        sheet_config = [
            ("🛡 Resilient Reversal", "Resilient Reversal — Held EMA stack during market weakness"),
        ],
        desc_map = {
            "🛡 Resilient Reversal": "Stocks that held their EMA stack while the broader market sold off — highest relative strength in downturns",
        }
    )

# ── MARKET HEALTH DASHBOARD ───────────────────────────────────
elif st.session_state.page == "market_health":
    render_generic_scanner(
        page_key     = "market_health",
        title        = "Market Health Dashboard",
        icon         = "🩺",
        accent_color = "#22c55e",
        file_prefix  = "Market_Health_",
        sheet_config = [
            ("🏆 9-9 Watchlist",   "9-9 Watchlist — All 9 Minervini conditions met"),
            ("👀 8-9 Monitor",     "8-9 Monitor — One condition away"),
            ("📈 RS Leaders Today","RS Leaders Today — Biggest RS movers"),
            ("📅 Daily Health Log","Daily Health Log — Breadth history"),
            ("🔍 Rally Quality",   "Rally Quality — FTD and rally tracker"),
            ("🚀 N500 Surge",      "N500 Surge — Nifty 500 outliers today"),
            ("📊 All Stocks",      "All Stocks — Full breadth universe"),
        ],
        desc_map = {
            "🏆 9-9 Watchlist":    "Every stock passing all 9 Minervini SEPA conditions today",
            "👀 8-9 Monitor":      "Stocks passing 8 of 9 conditions — one trigger away from ELITE",
            "📈 RS Leaders Today": "Stocks with biggest RS improvement today — emerging leaders",
            "📅 Daily Health Log": "Historical breadth readings — track market health over time",
            "🔍 Rally Quality":    "Follow-Through Day detection and rally quality assessment",
            "🚀 N500 Surge":       "Nifty 500 stocks surging significantly vs the index today",
            "📊 All Stocks":       "Full universe breadth data — all stocks with health metrics",
        }
    )

# ── RS OUTPERFORMER ───────────────────────────────────────────
elif st.session_state.page == "rs_outperformer":
    render_generic_scanner(
        page_key     = "rs_outperformer",
        title        = "RS Outperformer",
        icon         = "🥇",
        accent_color = "#f5a623",
        file_prefix  = "RS_Outperformer_",
        sheet_config = [
            ("Elite RS",  "Elite RS — Strongest outperformers vs Nifty"),
            ("Strong RS", "Strong RS — Solid outperformance"),
            ("Mild RS",   "Mild RS — Modest outperformance"),
            ("All RS",    "All RS — Complete ranked universe"),
            ("Summary",   "Summary — Period overview"),
        ],
        desc_map = {
            "Elite RS":  "Top-tier stocks that have significantly outperformed Nifty since the chosen start date",
            "Strong RS": "Strong outperformers — meaningful gap vs Nifty over the period",
            "Mild RS":   "Mild outperformers — slight edge over Nifty, worth watching",
            "All RS":    "Full universe ranked by RS outperformance vs Nifty",
            "Summary":   "Period summary — Nifty return, number of outperformers, top performers",
        }
    )
